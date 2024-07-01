import pandas as pd
import os
from openpyxl import load_workbook

def find_value_by_condition(excel_path, specified_variable,sheetname=0):
    # Load the Excel file
    df = pd.read_excel(excel_path,sheet_name=sheetname)
    
    # Initialize variable to store the previous value in the fourth column for comparison
    previous_value_in_fourth_column = None

    output= []

    # Loop through the DataFrame to find the matching condition
    for i in range(len(df)):
        # Check if the current row's third column matches the specified variable
        if df.iloc[i, 2] == specified_variable:
            # If this is not the first occurrence and the fifth column value has changed
            if previous_value_in_fourth_column is not None and df.iloc[i, 4] != previous_value_in_fourth_column:
                # Return the value in the first column of this row
                output.append((df.iloc[i, 2], df.iloc[i,3], df.iloc[i, 1], df.iloc[i, 0]))
            # Update the previous value in the fifth column for the next comparison
            previous_value_in_fourth_column = df.iloc[i, 4]
    
    # If no matching condition is found, return None
    return output

def first_instances(excel_path,sheet = 0):
    # Load the Excel file, add `header=0` if your file has headers, or remove it if it doesn't
    df = pd.read_excel(excel_path,sheet, header=0)
    
    # Optional: Check for and drop rows where any of the required columns have NaN values
    df = df.dropna(subset=[df.columns[2], df.columns[1], df.columns[0]])

    # Dictionary to keep track of first instances
    first_occurrences = {}

    # Iterate over each row in the DataFrame
    for index, row in df.iterrows():
        # Get the value in the third column
        key = row[df.columns[2]]  # Use column names if known, e.g., 'Column_Name'

        # If the value has not been recorded yet, add it to the dictionary
        if key not in first_occurrences and not pd.isna(key):
            # Store the tuple (value in column 3, value in column 2, value in column 1)
            first_occurrences[key] = (row[df.columns[2]],row[df.columns[4]], row[df.columns[1]], row[df.columns[0]])

    # Return a list of all recorded first occurrences
    return list(first_occurrences.values())

def unique_values_in_third_col(excel_path, sheet_name):
    # Load the Excel file
    df = pd.read_excel(excel_path, header=None, sheet_name=sheet_name)
    
    # Select the third row (indexing starts at 0, so index 2 is the third row)
    third_column = df.iloc[1:,2]
    
    # Get unique values from the third row
    unique_values = third_column.unique()
    
    # Return the unique values
    return unique_values

def get_exit_date(firm,excel_path,sheet):
    occurences = first_instances(excel_path,sheet)
    firm_index = [tup[0] for tup in occurences].index(firm)
    recent_year = occurences[0][-1]
    if occurences[firm_index][-1] == recent_year:
        return []
    return [occurences[firm_index]]

def total_exit_dates(firm, excel_path, sheet):
    return get_exit_date(firm, excel_path, sheet) + find_value_by_condition(excel_path, firm)

def all_exit_dates(excel_path,sheet):
    firms = unique_values_in_third_col(excel_path, sheet)
    exit_data = []
    for firm in firms:
        exit_data+=total_exit_dates(firm, excel_path, sheet)
    return exit_data

def new_all_exit_dates(excel_path,sheet):
    firms = unique_values_in_third_col(excel_path, sheet)
    exit_data = []
    for firm in firms:
        exit_data+=find_value_by_condition(excel_path, firm, sheet)
    return exit_data

def max_row_in_column(excel_path, sheet_name, column):
    # Load the workbook and sheet
    book = load_workbook(excel_path, data_only=True)
    if sheet_name not in book.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' does not exist in the workbook")
    sheet = book[sheet_name]

    # Initialize the max row variable
    max_row = 0

    # Iterate over the specified column
    for cell in sheet[column]:
        if cell.value is not None:
            max_row = cell.row

    return max_row

def export_tuples_to_excel(tuples, excel_path, sheet_name, start_col=0):
    # Create a Pandas Excel writer using openpyxl as the engine    
    # Ensure the file exists to avoid creating a new one by accident
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"No such file: '{excel_path}'")

    with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a',if_sheet_exists='overlay') as writer:
        # # Convert the list of tuples to a DataFrame
        if 'openpyxl' in writer.supported_extensions:
            writer.book = load_workbook(excel_path)
        
        # Check if the sheet exists, and find the next available row
        if sheet_name in writer.sheets:
            startrow=0
            for cell in writer.book[sheet_name][start_col]:
                if cell.value is not None:
                    startrow = cell.row
            # startrow = writer.sheets[sheet_name].max_row
        else:
            startrow = 0

        # Convert the list of tuples to a DataFrame
        df = pd.DataFrame(tuples)
        
        # Write DataFrame to Excel, appending without header and index
        df.to_excel(writer, sheet_name=sheet_name, startrow=startrow,
                    startcol=start_col, index=False, header=False)

def write_excel(raw_data, excel_path, raw_sheet, new_sheet, start_col):
    exit_data = all_exit_dates(raw_data, raw_sheet)
    export_tuples_to_excel(exit_data, excel_path, new_sheet, start_col)

def new_write_excel(raw_data, excel_path, raw_sheet, new_sheet, start_col):
    exit_data = new_all_exit_dates(raw_data, raw_sheet)
    export_tuples_to_excel(exit_data, excel_path, new_sheet, start_col)

# Example usage:
raw_data = "/Users/lucasg17/Documents/GitHub/Health-Innovation/Master.xlsx"
excel_path = "/Users/lucasg17/Downloads/Raw Exit Data (1).xlsx"
testing = '/Users/lucasg17/Documents/GitHub/Health-Innovation/Testing.xlsx'
sheet_name = 0
firm = 'BICYCLE THERAPEUTICS PLC'
write_excel(raw_data, testing, 'SV Health Investors, LLC', 'Master_SV1', 1)